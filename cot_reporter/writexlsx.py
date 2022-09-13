import glob
import pandas as pd
from openpyxl import load_workbook

# Convert archive file(s) to workbook
for f in glob.glob("downloaded/f_*.txt"):
    read_file = pd.read_csv(f)
    read_file.to_excel("stage/f_year.xlsx", index = None, header=True)

# Load workbook
wb = load_workbook('stage/f_year.xlsx')

# Define workbook sheets
sheet1 = wb.active
sheet2 = wb.create_sheet("Mysheet")
sheet3 = wb.create_sheet("Mysheet",0)
sheet1.title = "rawData"
sheet2.title = "bleesData"

# Copy source columns to target sheet
for i in range(1, sheet1.max_row + 1):
    sheet2["A" + str(i)].value = sheet1["A" + str(i)].value
    sheet2["B" + str(i)].value = sheet1["B" + str(i)].value
    sheet2["C" + str(i)].value = sheet1["I" + str(i)].value
    sheet2["D" + str(i)].value = sheet1["J" + str(i)].value
    sheet2["E" + str(i)].value = sheet1["K" + str(i)].value
    sheet2["F" + str(i)].value = sheet1["L" + str(i)].value
    sheet2["G" + str(i)].value = sheet1["M" + str(i)].value
    sheet2["H" + str(i)].value = sheet1["N" + str(i)].value
    sheet2["I" + str(i)].value = sheet1["O" + str(i)].value
    sheet2["J" + str(i)].value = sheet1["Q" + str(i)].value
    sheet2["K" + str(i)].value = sheet1["R" + str(i)].value

# Perform calculations in new sheet
sheet3["A1"] = "Report_Date"

# Save workbook and close
wb.save("f_year.xlsx")